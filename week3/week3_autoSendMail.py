from NaverNewsCrawler import NaverNewsCrawler
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import smtplib
import re
from openpyxl import load_workbook

# 단계를 1 ~ 5 단계로 나누어 기재하였습니다.
# 1, 2 단계는 제공 받은 템플릿 입니다.

# 1.
# [제공된 템플릿] 네이버뉴스 크롤러 를 이용한 추출 및 엑셀 저장
keywords = input("검색할 키워드를 입력하세요: ")
filename = input("결과를 담을 파일명을 입력하세요: ")

crawler = NaverNewsCrawler(keywords)
crawler.get_news(filename + '.xlsx')

# 2.
# [제공된 템플릿] 인터넷 강의를 통한 메일 발송 함수
def send_mail(name, addr, subject, contents, attachment=None):
    # 이메일 유효성 확인
    if not re.match('(^[a-zA-Z0-9_.-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)', addr):
        print('Wrong email')
        return

    msg = MIMEMultipart('alternative') # smtp 서버가 원하는 형태의 text 메일을 포함하겠다
    if attachment:
        msg = MIMEMultipart('mixed') # 만드는 메일의 내용이 text 이외의 것도 있다

    msg['From'] = SMTP_USER
    msg['To'] = addr
    msg['Subject'] = name + '님, ' + subject

    text = MIMEText(contents, _charset='utf-8')
    msg.attach(text)

    # 첨부파일이 있다면
    if attachment:
        from email.mime.base import MIMEBase
        from email import encoders

        # 첨부파일 인코딩
        file_data = MIMEBase('application', 'octect-stream')
        file_data.set_payload(open(attachment, 'rb').read())
        encoders.encode_base64(file_data)

        # 첨부파일명 지정
        import os
        filename = os.path.basename(attachment)
        file_data.add_header('Content-Disposition', 'attachment; filename="'+filename+'"')
        msg.attach(file_data)
    
    # 서버에 정보 전송 및 요청
    smtp = smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT)
    smtp.login(SMTP_USER, SMTP_PASSWORD)
    smtp.sendmail(SMTP_USER, addr, msg.as_string())
    smtp.close()

# 3.
# 발송자 정보
SMTP_SERVER = 'smtp.gmail.com'
SMTP_PORT = 465
SMTP_USER = 'iamidlek@gmail.com'
SMTP_PASSWORD = '****' # 제출을 위해 * 로 임시작성

# 4.
# 필요한 파일 열기 및 가져오기(수신자 정보, 크롤링 정보)
data = load_workbook('email list.xlsx', read_only=True)
receiver = data.active

openfile = load_workbook(filename, read_only=True)
readfile = openfile.active

# 5.
# 자동 발송 반복문(정보 설정 및 함수 호출)
# (수취인 1명당 * 뉴스 10건씩 발송)
for i in range(3, receiver.max_row+1):
  if receiver[i][1].value == None:
    continue
  rname = receiver[i][1].value
  raddr = receiver[i][2].value
  for i in range(2, readfile.max_row+1):
    if readfile[i][1].value == None:
      continue
    sub = readfile[i][1].value
    con = readfile[i][3].value
    send_mail(rname,raddr,sub,con)
